from manager.models import MTDTA_UPLOADER
from manager.models import MTDTA_UPLOADER_COLS
from manager.models import UTL_FILE_LOADER_BOT_LOGS

from django.core.exceptions import FieldError

from file_loader.models import *

from manager.uploader import UploadLogic

from django.apps import apps
from django.utils import timezone

import datetime
import csv
from openpyxl import load_workbook
from django.db import transaction
import ast
import os
import re
from io import TextIOWrapper

class InsertLogic:

    # Insert data into database
    def properInsert(self, inputFile, uploaderMetadata, uploaderMetadataColumns, uploadType):
        returned = []
        valid = True
        errors = []
        warnings = []
        responses = []
        required = []
        maxLength = []
        names = []

        startRow = uploaderMetadata[15]
        uploaderName = uploaderMetadata[1]
        targetSchemaName = uploaderMetadata[7]
        targetTableName = re.sub(r'[\W_]', '', uploaderMetadata[8])
        targetTable = apps.get_model('file_loader', targetTableName)
        fileDate = str(datetime.datetime.now())
        fileName = inputFile.name

        # Get max length, required, and names from metadata columns
        for metaCol in uploaderMetadataColumns:
            required.append(metaCol[4])
            maxLength.append(metaCol[7])
            names.append(metaCol[2])
        required.insert(0, 'Y')
        maxLength.insert(0, 255)
        names.insert(0, 'id')

        if(uploaderMetadata):

            # Read file and store into a variable
            # Max length validation is performed during file reading
            print("Read file")
            print(datetime.datetime.time(datetime.datetime.now()))

            # Reading for csv files
            if(uploaderMetadata[5].lower() == '.csv'):
                if(uploadType == 'auto'):
                    ws = csv.reader(inputFile, delimiter=uploaderMetadata[16])
                # TODO: Assisted upload of csv files is broken
                elif(uploadType == 'assisted'):
                    paramFile = TextIOWrapper(inputFile)
                    ws = csv.reader(paramFile, delimiter=uploaderMetadata[16])
                nCols = len(uploaderMetadataColumns)
                i = 0
                all = []
                try:
                    for row in ws:
                        r = []
                        j = 0
                        columnNumber = 1
                        exceedMaxLength = False
                        for cell in row:
                            if(cell != None and columnNumber <= nCols):
                                r.append(str(cell))
                                if(len(str(cell)) > int(maxLength[columnNumber])):
                                    exceedMaxLength = True
                                    maxLengthFound = len(str(cell))
                                    break
                            elif(cell.value == None and columnNumber <= nCols):
                                r.append(None)
                            columnNumber = columnNumber + 1
                        if(exceedMaxLength):
                            errors.append("Some values of file column '" + names[columnNumber] + "' exceeds maximum length. Expected: '" + str(maxLength[columnNumber]) + "', Found: " + str(maxLengthFound) + "'.")
                            valid = False
                            break
                        else:
                            for j in range(len(row), nCols):
                                r.append(None)
                                
                            # Add file name and file date data
                            r.append(fileName)
                            r.append(fileDate)
                            if(len(r) > 0):
                            # if(i > startRow):
                                r.insert(0,i)
                                all.append(r)
                        i = i + 1
                except Exception:
                    errors.append("Some values found in the file are not encoded in utf-8.")
                    valid = False
                if(uploadType == 'assisted'):
                    paramFile.detach()
                    
            # Reading for xls and xlsx files
            else:
                wb = load_workbook(inputFile, read_only=True)
                sheetNames = wb.get_sheet_names()
                sheetIndex = 0
                for s in sheetNames:
                    if(str(s) == uploaderMetadata[6]):
                        break
                    sheetIndex = sheetIndex + 1
                ws = wb[sheetNames[sheetIndex]]
                nCols = len(uploaderMetadataColumns)

                i = 0
                all = []
                for row in ws.iter_rows():
                    r = []
                    j = 0
                    columnNumber = 1
                    exceedMaxLength = False
                    for cell in row:
                        if(cell.value != None):
                            r.append(str(cell.value))
                            if(len(str(cell.value)) > int(maxLength[columnNumber])):
                                exceedMaxLength = True
                                maxLengthFound = len(str(cell.value))
                                break
                        elif(cell.value == None and columnNumber <= nCols):
                            r.append(None)
                        columnNumber = columnNumber + 1
                    if(exceedMaxLength):
                        errors.append("Some values of file column '" + names[columnNumber] + "' exceeds maximum length. Expected: '" + str(maxLength[columnNumber]) + "', Found: " + str(maxLengthFound) + "'.")
                        valid = False
                        break
                    else:
                        for j in range(len(row), nCols):
                            r.append(None)

                        # TODO: file name and file date column validation
                        # Add file name and file date data
                        r.append(fileName)
                        r.append(fileDate)
                        if(len(r) > 0):
                            if(i > startRow):
                                r.insert(0,i)
                                all.append(r)
                    i = i + 1
            # Insert to database
            if(valid):        
                print("Insert to database")
                print(datetime.datetime.time(datetime.datetime.now()))
                try:
                    with transaction.atomic():
                        targetTable.objects.using(targetSchemaName).all().delete()
                        for r in all:
                            t = targetTable(*r)
                            t.save(using=targetSchemaName)
                except Exception:
                    errors.append("Insert to database failed. Possible issues: (1) Values on file exceeds maximum allowed length; (2) Required columns does not exist in the target table: [id], [file_name], [file_date]; (3) File number of columns is not equal to target table number of columns")

            # Blank values per required column validation
            # print("Blank values validation")
            # print(datetime.datetime.time(datetime.datetime.now()))
            # if(valid):
            #     columnNumber = 0
            #     for name in names:
            #         values = []
            #         if(required[columnNumber] == 'Y'):
            #             hasBlank = False
            #             name = re.sub(r'[\W ]', '_', name)
            #             name = name.lower()
            #             try:
            #                 values = targetTable.objects.using(targetSchemaName).values_list(name, flat=True)   
            #                 for value in values:
            #                     if(value == None or value == ''):
            #                         hasBlank = True
            #                         break
            #                 if(hasBlank and required[columnNumber] == 'Y'):
            #                     warnings.append("Column has blank values: '" + names[columnNumber] + "'. This column is required.")
            #             except Exception:
            #                 valid = False
            #                 errors.append("Field error encountered. Some column names are invalid due to special characters.")
            #         columnNumber = columnNumber + 1
            
        else:
            valid = False
        print(datetime.datetime.time(datetime.datetime.now()))
        returned.append(None)
        returned.append(valid)
        returned.append(errors)
        returned.append(warnings)
        return returned

    # Truncate table
    def truncateTable(self, uploaderMetadata):
        targetSchemaName = uploaderMetadata[7]
        targetTableName = re.sub(r'[\W_]', '', uploaderMetadata[8])
        targetTable = apps.get_model('file_loader', targetTableName)

        try:
            targetTable = apps.get_model('file_loader', targetTableName)
        except Exception:
            valid = False
        if(targetTable):
            try:
                targetTable.objects.using(targetSchemaName).all().delete()
            except Exception:
                valid = False

    def insertLog(uploaderMetadataRaw, fileFullPath, errors, warnings, sendEmail, truncate, startTimeStamp):
        returned = []
        endTimeStamp = timezone.now()
        status = None
        errs = None
        warns = None

        if(errors):
            status = 'Failure'
            errs = str(errors)
        else:
            status = 'Success'
            errs = None
            if(warnings):
                warns = str(warnings)
            else:
                warns = None

        # try:
        with transaction.atomic():
            t = UTL_FILE_LOADER_BOT_LOGS(
                uploader_name = uploaderMetadataRaw[1], 
                last_update_uid = uploaderMetadataRaw[9], 
                update_start_timestamp = startTimeStamp,
                update_end_timestamp = endTimeStamp,
                status = status,
                error_details = errs,
                warning_details = warns,
                input_file = fileFullPath,
                email_sent = sendEmail, 
                table_truncated = truncate
            )
            t.save()
        # except Exception:
        #     errors.append("Insert to log table failed. Possible issues: (1) Table UTL_FILE_LOADER_BOT_LOGS not in the database; ")

        returned.append(errors)

        return returned