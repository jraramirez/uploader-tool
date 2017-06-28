from shift.models import TestFIN005Raw
from shift.models import META_SHIFT

from openpyxl import load_workbook
from django.db import transaction
import ast

class UploadLogic:

    # Read the metadata of the file
    def readFileMetadata(inputFile):
        wb = load_workbook(inputFile)
        contents = []

        sheetNames = wb.get_sheet_names()       # Sheet Names
        fileName = inputFile.name               # File Name
        fileType = inputFile.content_type       # File Type
        colNames = []                           # Column Names and nCols
        nCols = 0
        colNamesLength = 0
        ws = wb[sheetNames[0]]
        for row in ws.rows:
            for cell in row:
                colNames.append(cell.value)
                nCols = nCols + 1
            break

        contents = [sheetNames, colNames, fileName, fileType, nCols]
        return contents

    # Get the metadata from database
    def getMetadata():
        metadata = []
        if(META_SHIFT.objects.all().exists()):
            metadata = META_SHIFT.objects.all()
            metadata = [ast.literal_eval(metadata[0].sheetNames), metadata[0].colNames, metadata[0].fileName, metadata[0].fileType, metadata[0].numberOfColumns]
        return metadata

    # Get the labels from database
    # To do: Get labels from database
    def getLabels():
        labels = ['Sheet Names', 'Column Names', 'File Name', 'File Type', 'Number of Columns']
        return labels

    # Read the metadata of the file
    def validateFileMetadata(inputFile):
        valid = True
        errors = []
        wb = load_workbook(inputFile)
        contents = []
        metadata = []

        sheetNames = wb.get_sheet_names()       # Sheet Names
        fileName = inputFile.name               # File Name
        fileType = inputFile.content_type       # File Type
        colNames = []                           # Column Names and nCols
        nCols = 0
        ws = wb[sheetNames[0]]
        for row in ws.rows:
            for cell in row:
                colNames.append(cell.value)
                nCols = nCols + 1
            break
        
        if(not META_SHIFT.objects.all().exists()):
            valid = False
            errors.append("No metadata saved in the database. You must first save a metadata in the database.\n")
        
        else:
            metadata = META_SHIFT.objects.all()

            # Validate sheet names
            metadataSheetNames = ast.literal_eval(metadata[0].sheetNames)
            for sheet, metaSheet in zip(sheetNames, metadataSheetNames):
                if(sheet != metaSheet):
                    errors.append("Invalid sheet name(s):" + "Expected:" + metaSheet + "Found: " + sheet)
                    valid = False

            # Validate number of columns and column names
            if(nCols == metadata[0].numberOfColumns):
                metaColNames = ast.literal_eval(metadata[0].colNames)
                for col, metaCol in zip(colNames, metaColNames):
                    if(col != metaCol):
                        errors.append("Column name mismatch:" + " Expected:" + metaCol + " Found: " + col)
                        valid = False
            else:
                errors.append("Invalid number of columns: " + "Expected: " + metadata[0].numberOfColumns + " Found: " + nCols)
                valid = False

            # Validate file type
            if(fileType != metadata[0].fileType):
                errors.append("Invalid file type:\n" + "Expected:" + metadata[0].fileType +"\tFound:" + fileType)
                valid = False

        errors.insert(0, valid)
        return errors

    # Read the metadata of the file
    def getFileMetadataChanges(inputFile):
        valid = True
        hasChanges = False
        changes = []
        wb = load_workbook(inputFile)
        contents = []

        sheetNames = wb.get_sheet_names()       # Sheet Names
        fileName = inputFile.name               # File Name
        fileType = inputFile.content_type       # File Type
        colNames = []                           # Column Names and nCols
        nCols = 0
        ws = wb[sheetNames[0]]
        for row in ws.rows:
            for cell in row:
                colNames.append(cell.value)
                nCols = nCols + 1
            break
        if(not META_SHIFT.objects.all().exists()):
            valid = True
            hasChanges = True
            changes.append("No metadata saved in the database. This is the first time a metadata will be saved for this upload.\n")

        else:
            metadata = META_SHIFT.objects.all()
            # Validate sheet names
            metadataSheetNames = ast.literal_eval(metadata[0].sheetNames)
            for sheet, metaSheet in zip(sheetNames, metadataSheetNames):
                if(sheet != metaSheet):
                    changes.append("Sheet name(s) will changed from: " + metaSheet + " to " + sheet)
                    hasChanges = True

            # Validate number of columns and column names
            if(nCols == metadata[0].numberOfColumns):
                metaColNames = ast.literal_eval(metadata[0].colNames)
                for col, metaCol in zip(colNames, metaColNames):
                    if(col != metaCol):
                        changes.append("Column name will be changed from " + metaCol + " to " + col)
                        hasChanges = True
            else:
                changes.append("Cannot change the number of columns: " + "Expected: " + metadata[0].numberOfColumns + " Found: " + nCols)
                hasChanges = True
                valid = False

            # Validate file type
            if(fileType != metadata[0].fileType):
                changes.append("File type will be changed from " + metadata[0].fileType +" to " + fileType)
                hasChanges = True

        changes.insert(0, valid)
        changes.insert(0, hasChanges)
        return changes

    # Read and save the metadata of the file
    # To do: save labels
    def saveFileMetadata(inputFile):
        wb = load_workbook(inputFile)
        contents = []

        sheetNames = wb.get_sheet_names()       # Sheet Names
        fileName = inputFile.name               # File Name
        fileType = inputFile.content_type       # File Type
        colNames = []                           # Column Names and nCols
        nCols = 0
        ws = wb[sheetNames[0]]
        for row in ws.rows:
            for cell in row:
                colNames.append(cell.value)
                nCols = nCols + 1
            break

        # Save Metadata
        META_SHIFT.objects.all().delete()
        m = META_SHIFT(1, sheetNames, colNames, fileName, fileType, nCols)
        m.save()

        contents = [sheetNames, colNames, fileName, fileType, nCols]
        return contents

    # Insert data into database
    def properInsert(inputFile):
        TestFIN005Raw.objects.all().delete()

        with transaction.atomic():
            wb = load_workbook(inputFile)
            ws = wb['P&G Raw']

            i = 0
            for row in ws.iter_rows():
                r = []
                j = 0
                for cell in row:
                    if(cell.value != None):
                        r.append(str(cell.value))
                    else:
                        r.append('')
                for j in range(len(row), 32):
                    r.append('')
                if(len(r) > 0):
                    if(i > 0):
                        t = TestFIN005Raw(
                            i, r[0],r[1],r[2],r[3],r[4],r[5], '',
                            r[6],r[7],r[8],r[9],r[10],r[11],
                            r[12],r[13],r[14],r[15],r[16],
                            r[17],r[18],r[19],r[20],r[21],
                            r[22],r[23],r[24],r[25],r[26],
                            r[27],r[28],r[29],
                        )
                        t.save()
                    i = i + 1

